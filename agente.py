"""
agente.py
---------
Agente Buchholz - Mantenimiento Predictivo de subestaciones electricas.

Flujo:
  1. Lee el Excel 'Mantenimiento_Subestaciones.xlsx' (hojas Mediciones y Limites).
  2. Construye el arbol de objetos: Subestacion -> Equipo -> Visita -> Medicion.
  3. Genera un reporte legible (resumen, hallazgos, estado por equipo, historico).
  4. Si hay hallazgos, envia alerta por Telegram y reporte completo por correo.

Autor: Jose David Perez Munguia - UTH 2026.4
"""

import os
import json
import html
import datetime
import unicodedata
from openpyxl import load_workbook

from modelos import Medicion, Visita, Subestacion

# Ruta del Excel: misma carpeta que este archivo
CARPETA = os.path.dirname(os.path.abspath(__file__))
ARCHIVO = os.path.join(CARPETA, "Mantenimiento_Subestaciones.xlsx")
# Archivo donde se recuerda que alertas ya se conocian (para notificar solo cambios)
ESTADO_ALERTAS = os.path.join(CARPETA, "estado_alertas.json")

# Todos los parametros que se miden: (columna en Excel, nombre, unidad)
COLUMNAS = [
    ("Voltaje VDC (V)",            "Voltaje VDC",        "V"),
    ("Voltaje AC (V)",             "Voltaje AC",         "V"),
    ("Resistencia tierra (Ohm)",   "Resistencia tierra", "Ohm"),
    ("Corriente de fuga (mA)",     "Corriente de fuga",  "mA"),
    ("Corriente Fase A (A)",       "Corriente Fase A",   "A"),
    ("Corriente Fase B (A)",       "Corriente Fase B",   "A"),
    ("Corriente Fase C (A)",       "Corriente Fase C",   "A"),
    ("Potencia Activa (MW)",       "Potencia Activa",    "MW"),
    ("Potencia Reactiva (MVAR)",   "Potencia Reactiva",  "MVAR"),
    ("Num Operaciones",            "Num Operaciones",    "oper"),
]


# Ficha tecnica del equipo (datos de placa). Orden y etiqueta que se muestra:
FICHA_CAMPOS = [
    ("serie",     "Serie"),
    ("marca",     "Marca"),
    ("modelo",    "Modelo"),
    ("tipo",      "Tipo"),
    ("i_nominal", "I nominal (A)"),
    ("tension",   "Tension"),
    ("ano",       "Ano"),
    ("icc",       "ICC"),
]
# Para reconocer la etiqueta en el Excel (normalizada) -> clave interna.
FICHA_LABELS = {
    "SERIE": "serie", "MARCA": "marca", "MODELO": "modelo", "TIPO": "tipo",
    "I NOMINAL (A)": "i_nominal", "I NOMINAL": "i_nominal",
    "TENSION": "tension", "ANO": "ano", "ICC": "icc",
}


def _norm(texto):
    """Quita acentos, espacios y ':' finales, y pasa a mayusculas (para comparar)."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return t.strip().rstrip(":").strip().upper()


def _valor_derecha(fila, i):
    """Primer valor no vacio a la derecha de la celda i (saltando vacias)."""
    for j in range(i + 1, len(fila)):
        v = fila[j].value
        if v is None or str(v).strip() == "":
            continue
        if _norm(v) in FICHA_LABELS:        # choco con otra etiqueta -> sin valor
            return None
        return v
    return None


# ============================================================
# Lectura del Excel -> arbol de objetos
# ============================================================
def _fila_encabezado(hoja):
    """Devuelve el numero de fila donde empieza la tabla de visitas (la de 'Fecha')."""
    for idx, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
        if fila and _norm(fila[0]) == "FECHA":
            return idx
    return 1                                 # por si la hoja no tiene ficha arriba


def _leer_ficha(hoja, fila_header):
    """Lee la ficha tecnica (datos de placa) de las filas por encima del encabezado."""
    ficha = {}
    for fila in hoja.iter_rows(min_row=1, max_row=max(1, fila_header - 1)):
        for i, celda in enumerate(fila):
            clave = FICHA_LABELS.get(_norm(celda.value))
            if clave and clave not in ficha:
                valor = _valor_derecha(fila, i)
                if valor is not None and str(valor).strip() != "":
                    ficha[clave] = valor
    return ficha


def cargar_limites(wb):
    """Devuelve {parametro: {criterio, limite}} desde la hoja Limites."""
    ws = wb["Limites"]
    limites = {}
    for fila in ws.iter_rows(min_row=2, values_only=True):
        parametro, criterio, limite = (list(fila) + [None]*3)[:3]
        if parametro:
            limites[parametro] = {"criterio": criterio, "limite": limite}
    return limites


def _agregar_medicion(visita, fila, limites):
    """Llena una visita con las mediciones de una fila (dict columna->valor)."""
    for columna, parametro, unidad in COLUMNAS:
        lim = limites.get(parametro)
        if lim:
            visita.agregar(Medicion(parametro, fila.get(columna), unidad,
                                    lim["criterio"], lim["limite"]))
        else:
            visita.agregar(Medicion(parametro, fila.get(columna), unidad))


def construir_modelo(wb):
    """
    Arma el arbol Subestacion -> Equipo -> Visita -> Medicion leyendo:
      - la hoja indice 'Equipos' (Subestacion, Codigo, Tipo), y
      - una hoja por equipo (nombrada con su codigo) con sus visitas.
    Asi cada equipo tiene su propia tabla y los datos no se mezclan.
    """
    limites = cargar_limites(wb)
    subestaciones = {}                      # nombre -> Subestacion

    indice = wb["Equipos"]
    for fila in indice.iter_rows(min_row=2, values_only=True):
        nombre_sub, codigo, tipo = (list(fila) + [None] * 3)[:3]
        if not codigo:                      # fila vacia en el indice
            continue
        if codigo not in wb.sheetnames:     # equipo listado pero sin hoja propia
            print(f"  (aviso) El equipo {codigo} no tiene hoja de datos; se omite.")
            continue

        sub = subestaciones.setdefault(nombre_sub, Subestacion(nombre_sub))
        equipo = sub.obtener_equipo(codigo, tipo)

        # leer la hoja propia del equipo
        hoja = wb[codigo]
        fila_header = _fila_encabezado(hoja)            # donde empieza la tabla de visitas
        equipo.ficha = _leer_ficha(hoja, fila_header)   # datos de placa (arriba del encabezado)
        titulos = [c.value for c in hoja[fila_header]]
        for fila_v in hoja.iter_rows(min_row=fila_header + 1, values_only=True):
            datos = dict(zip(titulos, fila_v))
            if datos.get("Fecha") is None:  # fila vacia
                continue
            visita = Visita(str(datos["Fecha"]), datos.get("Cuadrilla"),
                            datos.get("Observacion") or "")
            _agregar_medicion(visita, datos, limites)
            equipo.agregar_visita(visita)

    return subestaciones


# ============================================================
# Utilidades para recorrer el modelo
# ============================================================
def todos_los_equipos(subestaciones):
    return [e for s in subestaciones.values() for e in s.equipos.values()]


def todos_los_hallazgos(subestaciones):
    """Lista de (equipo, visita, medicion) que estan en REVISAR."""
    salida = []
    for e in todos_los_equipos(subestaciones):
        for v in e.visitas:
            for m in v.hallazgos():
                salida.append((e, v, m))
    return salida


# ============================================================
# Reporte COMPLETO (para correo y consola)
# ============================================================
SEP = "=" * 60
SUB = "-" * 60


def _linea_param(visita, parametro, etiqueta):
    """Formatea una linea legible de una medicion: etiqueta ... valor [estado]."""
    m = visita.medicion(parametro)
    if m is None or m.valor is None:
        return f"        {etiqueta:.<24} (sin dato)"
    valor = f"{m.valor} {m.unidad}"
    if m.tiene_limite():
        return f"        {etiqueta:.<24} {valor:<11} [{m.clasificar()}]"
    return f"        {etiqueta:.<24} {valor}"


def reporte_completo(subestaciones, hoy):
    eq = todos_los_equipos(subestaciones)
    visitas = [v for e in eq for v in e.visitas]
    hallazgos = todos_los_hallazgos(subestaciones)
    L = []

    # ---- encabezado ----
    L.append(SEP)
    L.append("        AGENTE BUCHHOLZ - MANTENIMIENTO PREDICTIVO")
    L.append(f"        Reporte generado: {hoy}")
    L.append(SEP)
    L.append("")
    L.append(" RESUMEN")
    L.append(f"   Subestaciones : {len(subestaciones)}")
    L.append(f"   Equipos       : {len(eq)}")
    L.append(f"   Visitas       : {len(visitas)}")
    L.append(f"   Hallazgos     : {len(hallazgos)}  (requieren accion)")

    # ---- 1) hallazgos ----
    L.append("")
    L.append(SEP)
    L.append(" 1) HALLAZGOS QUE REQUIEREN ACCION")
    L.append(SEP)
    if not hallazgos:
        L.append("   Sin hallazgos: todos los parametros con limite estan NORMAL.")
    for e, v, m in hallazgos:
        L.append("")
        L.append(f"   [!] {e.codigo}  ({e.tipo})")
        L.append(f"       Subestacion : {e.subestacion}")
        L.append(f"       Fecha       : {v.fecha}        Cuadrilla: {v.cuadrilla}")
        L.append(f"       Parametro   : {m.parametro}")
        L.append(f"       Medido      : {m.valor} {m.unidad}   (limite {m.criterio}: {m.limite})")
        if v.observacion:
            L.append(f"       Observacion : {v.observacion}")

    # ---- 2) estado actual por equipo ----
    L.append("")
    L.append(SEP)
    L.append(" 2) ESTADO ACTUAL POR EQUIPO (ultima visita)")
    L.append(SEP)
    for e in eq:
        v = e.ultima_visita()
        if v is None:
            continue
        L.append("")
        L.append(f"   {e.codigo}  ({e.tipo})  -  {e.subestacion}")
        L.append(f"   Ultima visita: {v.fecha}   Cuadrilla: {v.cuadrilla}")
        L.append("     Voltajes:")
        L.append(_linea_param(v, "Voltaje VDC", "Voltaje VDC"))
        L.append(_linea_param(v, "Voltaje AC", "Voltaje AC"))
        L.append("     Puesta a tierra:")
        L.append(_linea_param(v, "Resistencia tierra", "Resistencia tierra"))
        L.append(_linea_param(v, "Corriente de fuga", "Corriente de fuga"))
        L.append("     Corrientes (monitoreo):")
        L.append(_linea_param(v, "Corriente Fase A", "Fase A"))
        L.append(_linea_param(v, "Corriente Fase B", "Fase B"))
        L.append(_linea_param(v, "Corriente Fase C", "Fase C"))
        L.append("     Potencias (monitoreo):")
        L.append(_linea_param(v, "Potencia Activa", "Activa"))
        L.append(_linea_param(v, "Potencia Reactiva", "Reactiva"))
        L.append("     Otros:")
        L.append(_linea_param(v, "Num Operaciones", "Num operaciones"))
        if v.observacion:
            L.append(f"     Observacion: {v.observacion}")
        L.append("   " + "." * 50)

    # ---- 3) historico de tierra ----
    L.append("")
    L.append(SEP)
    L.append(" 3) HISTORICO Y TENDENCIA - Resistencia de tierra (Ohm)")
    L.append(SEP)
    for e in eq:
        datos = e.historico("Resistencia tierra")
        if not datos:
            continue
        L.append("")
        L.append(f"   {e.codigo} ({e.tipo}) - {e.subestacion}")
        for fecha, valor in datos:
            estado = Medicion("Resistencia tierra", valor, "Ohm", "maximo", 5).clasificar()
            marca = "  <-- REVISAR" if estado == "REVISAR" else ""
            L.append(f"        {fecha}   {valor:>5} Ohm   {estado}{marca}")
        L.append(f"        Tendencia: {e.tendencia('Resistencia tierra')}")

    L.append("")
    L.append(SEP)
    return "\n".join(L)


# ============================================================
# Reporte CORTO (alerta para Telegram, con formato HTML)
# ============================================================
def _emoji_tendencia(texto):
    """Convierte el texto de tendencia ('subiendo (...)') en una flecha."""
    if texto.startswith("subiendo"):
        return "\U0001F4C8"          # grafico subiendo
    if texto.startswith("bajando"):
        return "\U0001F4C9"          # grafico bajando
    if texto.startswith("estable"):
        return "➡️"        # flecha a la derecha
    return ""


def _esc(texto):
    """Escapa <, > y & para que no rompan el formato HTML de Telegram."""
    return html.escape(str(texto))


def _url_dashboard():
    """
    URL publica del dashboard web, tomada de config.DASHBOARD_URL.
    Si no esta configurada (o esta vacia) devuelve '' y los mensajes no
    muestran ningun enlace. El dashboard es un archivo HTML; para tener un
    enlace que sirva a otras personas hay que publicarlo (GitHub Pages, etc.).
    """
    try:
        import config
        return (getattr(config, "DASHBOARD_URL", "") or "").strip()
    except Exception:
        return ""


def _enlace_telegram_dashboard():
    """Linea con el enlace al dashboard para Telegram (HTML); '' si no hay URL."""
    url = _url_dashboard()
    if not url:
        return ""
    return f'\U0001F310 <a href="{html.escape(url, quote=True)}">Ver dashboard web</a>'


def _boton_html_dashboard():
    """Boton/enlace al dashboard para los correos HTML; '' si no hay URL."""
    url = _url_dashboard()
    if not url:
        return ""
    u = html.escape(url, quote=True)
    return ('<p style="margin:18px 0;text-align:center;">'
            f'<a href="{u}" style="background:#0d6efd;color:#fff;'
            'text-decoration:none;padding:10px 20px;border-radius:6px;'
            'font-weight:bold;display:inline-block;">'
            '&#127760; Ver dashboard web</a></p>')


def reporte_corto(subestaciones, hoy):
    """
    Alerta breve para Telegram con formato HTML (emojis + negritas).
    Telegram HTML solo admite <b>, <i>, <u>, <code>, etc.; los saltos de
    linea son '\\n' normales y el texto dinamico se escapa con _esc().
    """
    hallazgos = todos_los_hallazgos(subestaciones)
    L = []
    L.append("\U0001F527 <b>AGENTE BUCHHOLZ - ALERTA</b>")
    L.append(f"\U0001F4C5 {hoy}")
    if hallazgos:
        L.append(f"⚠️ <b>{len(hallazgos)} hallazgo(s)</b> requieren accion")
    else:
        L.append("✅ <b>Sin hallazgos</b>: todos los parametros con limite estan NORMAL")

    for e, v, m in hallazgos:
        criterio = "max" if m.criterio == "maximo" else "min"
        flecha = _emoji_tendencia(e.tendencia(m.parametro))
        fecha_corta = str(v.fecha)[:10]
        L.append("")
        L.append(f"\U0001F534 <b>{_esc(e.codigo)}</b> <i>({_esc(e.tipo)})</i> - {_esc(e.subestacion)}")
        L.append(f"   • {_esc(m.parametro)}: <b>{m.valor} {_esc(m.unidad)}</b> "
                 f"(limite {criterio} {m.limite}) {flecha}".rstrip())
        L.append(f"   • \U0001F4CD Cuadrilla {_esc(v.cuadrilla)} · {fecha_corta}")
        if v.observacion:
            L.append(f"   • \U0001F4DD <i>{_esc(v.observacion)}</i>")

    if hallazgos:
        L.append("")
        L.append("\U0001F4E7 <i>Detalle completo enviado al correo.</i>")
    enlace = _enlace_telegram_dashboard()
    if enlace:
        L.append("")
        L.append(enlace)
    return "\n".join(L)


# ============================================================
# Reporte HTML SEGMENTADO (para correo): solo el equipo en alerta
# ============================================================
# Cada estado tiene su par de colores (fondo, texto):
_COLORES = {
    "REVISAR":   ("#f8d7da", "#842029"),   # rojo
    "NORMAL":    ("#d1e7dd", "#0f5132"),   # verde
    "MONITOREO": ("#e2e3e5", "#41464b"),   # gris
    "SIN DATO":  ("#fff3cd", "#664d03"),   # amarillo
}


def _badge(estado):
    """Devuelve una 'etiqueta' HTML coloreada segun el estado."""
    fondo, texto = _COLORES.get(estado, ("#e2e3e5", "#41464b"))
    return (f'<span style="background:{fondo};color:{texto};padding:2px 8px;'
            f'border-radius:10px;font-size:12px;font-weight:bold;">{estado}</span>')


def _celda(contenido, extra=""):
    """Una celda <td> con el estilo base de las tablas."""
    return f'<td style="padding:6px 10px;border:1px solid #dee2e6;{extra}">{contenido}</td>'


def reporte_html(subestaciones, hoy):
    """
    Genera el correo HTML ENFOCADO en los equipos que tienen alerta.
    No incluye los equipos que estan bien: solo el/los equipo(s) en alerta
    con su motivo y los datos de su ultima visita.

    Los graficos NO van en el correo (para mantenerlo ligero); el historico
    grafico se consulta en el dashboard web. Devuelve solo el HTML.
    """
    hallazgos = todos_los_hallazgos(subestaciones)

    # equipos unicos con alerta, conservando el orden de aparicion
    equipos_alerta = []
    for e, v, m in hallazgos:
        if e not in equipos_alerta:
            equipos_alerta.append(e)

    H = []
    H.append('<div style="font-family:Arial,Helvetica,sans-serif;color:#212529;'
             'max-width:760px;margin:auto;">')

    # ---- encabezado (rojo: es una alerta) ----
    H.append('<div style="background:#dc3545;color:#fff;padding:16px 20px;'
             'border-radius:8px 8px 0 0;">')
    H.append('<h2 style="margin:0;">&#9888;&#65039; Agente Buchholz &mdash; Alerta de Mantenimiento</h2>')
    H.append(f'<p style="margin:4px 0 0;opacity:.9;">Reporte generado: {hoy} '
             f'&middot; {len(equipos_alerta)} equipo(s) en alerta</p>')
    H.append('</div>')
    H.append('<div style="border:1px solid #dee2e6;border-top:none;padding:20px;'
             'border-radius:0 0 8px 8px;">')

    # ---- una seccion por equipo en alerta ----
    for e in equipos_alerta:
        suyos = [(vv, mm) for (ee, vv, mm) in hallazgos if ee is e]

        H.append('<h3 style="border-bottom:2px solid #dc3545;padding-bottom:4px;">'
                 f'&#128295; {_esc(e.codigo)} '
                 '<span style="font-weight:normal;color:#6c757d;font-size:15px;">'
                 f'({_esc(e.tipo)}) &mdash; {_esc(e.subestacion)}</span></h3>')

        # motivo de la alerta
        H.append('<div style="background:#f8d7da;color:#842029;padding:10px 14px;'
                 'border-radius:6px;margin-bottom:14px;">')
        H.append('<b>Motivo de la alerta:</b>'
                 '<ul style="margin:6px 0 0;padding-left:20px;">')
        for vv, mm in suyos:
            criterio = "max" if mm.criterio == "maximo" else "min"
            H.append(f'<li>{_esc(mm.parametro)}: <b>{mm.valor} {_esc(mm.unidad)}</b> '
                     f'(limite {criterio} {mm.limite}) &middot; '
                     f'visita {str(vv.fecha)[:10]}</li>')
        H.append('</ul></div>')

        # ultima visita completa
        v = e.ultima_visita()
        if v is not None:
            H.append(f'<p style="margin:0 0 4px;"><b>Ultima visita:</b> '
                     f'{str(v.fecha)[:10]} &middot; Cuadrilla {_esc(v.cuadrilla)}</p>')
            H.append('<table style="border-collapse:collapse;width:100%;'
                     'font-size:14px;margin-bottom:16px;">')
            for columna, parametro, unidad in COLUMNAS:
                m = v.medicion(parametro)
                if m is None:
                    continue
                valor = "(sin dato)" if m.valor is None else f"{m.valor} {m.unidad}"
                H.append('<tr>')
                H.append(_celda(_esc(parametro), "width:45%;"))
                H.append(_celda(valor))
                H.append(_celda(_badge(m.clasificar()), "text-align:center;width:120px;"))
                H.append('</tr>')
            H.append('</table>')
            if v.observacion:
                H.append('<p style="font-size:12px;color:#6c757d;margin:0 0 12px;">'
                         f'&#128221; {_esc(v.observacion)}</p>')

        H.append('<p style="font-size:12px;color:#6c757d;">Los graficos del historico '
                 '(corrientes, potencias y tierra) se consultan en el dashboard web.</p>')
        H.append('<hr style="border:none;border-top:1px solid #dee2e6;margin:24px 0;">')

    H.append(_boton_html_dashboard())
    H.append('<p style="font-size:12px;color:#adb5bd;margin-top:8px;">'
             'Generado automaticamente por el Agente Buchholz (mantenimiento predictivo) '
             '&middot; Jose David Perez Munguia &mdash; UTH 2026.4</p>')
    H.append('</div></div>')
    return "\n".join(H)


# ============================================================
# Notificar SOLO cuando algo cambie (alertas nuevas o resueltas)
# ============================================================
def _cargar_estado_alertas():
    """Conjunto de claves de alerta ('CODIGO|parametro') que ya se conocian."""
    try:
        with open(ESTADO_ALERTAS, encoding="utf-8") as f:
            return set(json.load(f))
    except (FileNotFoundError, ValueError, OSError):
        return set()


def _guardar_estado_alertas(claves):
    with open(ESTADO_ALERTAS, "w", encoding="utf-8") as f:
        json.dump(sorted(claves), f, ensure_ascii=False, indent=2)


def _texto_telegram_cambio(actuales, nuevas, resueltas, hoy):
    L = ["\U0001F514 <b>Cambios en alertas</b>", f"\U0001F4C5 {hoy}"]
    if nuevas:
        L.append("")
        L.append(f"\U0001F534 <b>Nuevas ({len(nuevas)}):</b>")
        for k in nuevas:
            e, v, m = actuales[k]
            crit = "max" if m.criterio == "maximo" else "min"
            L.append(f"  • <b>{_esc(e.codigo)}</b> ({_esc(e.subestacion)}): "
                     f"{_esc(m.parametro)} = <b>{m.valor} {_esc(m.unidad)}</b> "
                     f"(limite {crit} {m.limite})")
    if resueltas:
        L.append("")
        L.append(f"✅ <b>Resueltas ({len(resueltas)}):</b>")
        for k in resueltas:
            cod, par = k.split("|", 1)
            L.append(f"  • <b>{_esc(cod)}</b>: {_esc(par)}")
    enlace = _enlace_telegram_dashboard()
    if enlace:
        L.append("")
        L.append(enlace)
    return "\n".join(L)


def _texto_cambio(actuales, nuevas, resueltas, hoy):
    L = [f"Cambios en alertas - {hoy}", ""]
    if nuevas:
        L.append("NUEVAS:")
        for k in nuevas:
            e, v, m = actuales[k]
            L.append(f"  {e.codigo} ({e.subestacion}): {m.parametro} = "
                     f"{m.valor} {m.unidad} (limite {m.limite})")
    if resueltas:
        L.append("RESUELTAS:")
        for k in resueltas:
            cod, par = k.split("|", 1)
            L.append(f"  {cod}: {par}")
    return "\n".join(L)


def _html_cambio(actuales, nuevas, resueltas, hoy):
    H = ['<div style="font-family:Arial,Helvetica,sans-serif;color:#212529;'
         'max-width:680px;margin:auto;">']
    H.append('<div style="background:#0d6efd;color:#fff;padding:14px 18px;'
             'border-radius:8px 8px 0 0;"><h2 style="margin:0;">&#128276; '
             f'Agente Buchholz &mdash; Cambios en alertas</h2>'
             f'<p style="margin:4px 0 0;opacity:.9;">{hoy}</p></div>')
    H.append('<div style="border:1px solid #dee2e6;border-top:none;padding:18px;'
             'border-radius:0 0 8px 8px;">')
    if nuevas:
        H.append('<h3 style="color:#842029;border-bottom:2px solid #dc3545;'
                 'padding-bottom:4px;">&#128308; Nuevas alertas</h3>')
        H.append('<table style="border-collapse:collapse;width:100%;font-size:14px;'
                 'margin-bottom:16px;"><tr style="background:#f8f9fa;">')
        for c in ["Equipo", "Subestacion", "Parametro", "Medido", "Limite", "Fecha"]:
            H.append(f'<th style="padding:6px 10px;border:1px solid #dee2e6;'
                     f'text-align:left;">{c}</th>')
        H.append('</tr>')
        for k in nuevas:
            e, v, m = actuales[k]
            crit = "max" if m.criterio == "maximo" else "min"
            H.append('<tr style="background:#fff5f5;">'
                     + _celda(f'<b>{_esc(e.codigo)}</b>') + _celda(_esc(e.subestacion))
                     + _celda(_esc(m.parametro))
                     + _celda(f'<b style="color:#842029;">{m.valor} {_esc(m.unidad)}</b>')
                     + _celda(f'{crit} {m.limite}') + _celda(str(v.fecha)[:10]) + '</tr>')
        H.append('</table>')
    if resueltas:
        H.append('<h3 style="color:#0f5132;border-bottom:2px solid #198754;'
                 'padding-bottom:4px;">&#9989; Resueltas</h3><ul>')
        for k in resueltas:
            cod, par = k.split("|", 1)
            H.append(f'<li><b>{_esc(cod)}</b>: {_esc(par)}</li>')
        H.append('</ul>')
    H.append(_boton_html_dashboard())
    H.append('<p style="font-size:12px;color:#adb5bd;margin-top:14px;">Notificacion '
             'automatica de cambios &middot; consulta el dashboard para el detalle.</p>')
    H.append('</div></div>')
    return "\n".join(H)


def notificar_si_cambio(subestaciones, hoy, enviar=True):
    """
    Compara las alertas actuales con las ya conocidas (estado_alertas.json) y
    notifica SOLO si hay alertas nuevas o resueltas. Devuelve True si hubo cambio.
    Pensado para la revision automatica (no satura: si nada cambia, no manda nada).
    """
    hallazgos = todos_los_hallazgos(subestaciones)
    actuales = {f"{e.codigo}|{m.parametro}": (e, v, m) for e, v, m in hallazgos}
    previas = _cargar_estado_alertas()
    nuevas = [k for k in actuales if k not in previas]
    resueltas = [k for k in previas if k not in actuales]

    if not nuevas and not resueltas:
        print("Sin cambios en las alertas; no se notifica.")
        return False

    print(f"Cambios: {len(nuevas)} nueva(s), {len(resueltas)} resuelta(s).")
    if enviar:
        try:
            import notificaciones
            notificaciones.notificar(
                "Agente Buchholz - Cambios en alertas",
                _texto_cambio(actuales, nuevas, resueltas, hoy),
                _texto_telegram_cambio(actuales, nuevas, resueltas, hoy),
                parse_mode_telegram="HTML",
                cuerpo_html_correo=_html_cambio(actuales, nuevas, resueltas, hoy),
            )
        except Exception as e:
            print("  -> No se pudo notificar:", e)
    _guardar_estado_alertas(actuales.keys())
    return True


# ============================================================
# Programa principal
# ============================================================
def main():
    if not os.path.exists(ARCHIVO):
        print("No encuentro el Excel en:", ARCHIVO)
        return

    hoy = datetime.date.today().isoformat()
    wb = load_workbook(ARCHIVO, data_only=True)
    subestaciones = construir_modelo(wb)

    completo = reporte_completo(subestaciones, hoy)
    print(completo)

    # Notificar solo si hay algo que revisar
    if todos_los_hallazgos(subestaciones):
        try:
            import notificaciones
            print("\nEnviando notificaciones...")
            notificaciones.notificar(
                "Agente Buchholz - Hallazgos",
                completo,                                  # correo: texto plano (respaldo)
                reporte_corto(subestaciones, hoy),         # telegram: alerta corta
                parse_mode_telegram="HTML",                # telegram con formato
                cuerpo_html_correo=reporte_html(subestaciones, hoy),   # correo: HTML (sin graficos)
            )
        except ImportError:
            print("\n(Notificaciones aun no configuradas: falta config.py)")
        except Exception as e:
            print("\n(No se pudo notificar:", e, ")")


if __name__ == "__main__":
    main()
