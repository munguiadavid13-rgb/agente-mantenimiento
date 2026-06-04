"""
agregar_equipo.py
-----------------
Agrega un equipo nuevo CORRECTAMENTE: crea a la vez la fila en la hoja
'Equipos' (indice) Y su hoja de datos con las columnas estandar. Asi nunca
se desajustan el indice y las hojas.

Uso (dos formas):
  1) Con argumentos:
       python agregar_equipo.py "Coyoles Central" "CCE-32L43" "Restaurador"
  2) Sin argumentos (te pregunta paso a paso):
       python agregar_equipo.py

Despues de agregarlo, abre el Excel y llena las visitas en la hoja del equipo.

Autor: Jose David Perez Munguia - UTH 2026.4
"""

import sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

import agente

CUADRILLAS = "LCB1,LCB2,TCO1"


def _headers():
    """Columnas estandar de una hoja de equipo (fuente unica: agente.COLUMNAS)."""
    return ["Fecha", "Cuadrilla"] + [col for col, _, _ in agente.COLUMNAS] + ["Observacion"]


def agregar_equipo(subestacion, codigo, tipo, archivo=None):
    """Crea la fila en 'Equipos' y la hoja de datos del equipo. Reordena hojas."""
    archivo = archivo or agente.ARCHIVO
    wb = load_workbook(archivo)            # sin data_only: conserva formato/validaciones
    neg = Font(bold=True)

    indice = wb["Equipos"]
    codigos = {r[1].value for r in indice.iter_rows(min_row=2) if r[1].value}
    if codigo in codigos or codigo in wb.sheetnames:
        print(f"El equipo '{codigo}' YA existe. No se hace nada.")
        return False

    # 1) fila en el indice
    indice.append([subestacion, codigo, tipo])

    # 2) hoja de datos: ficha tecnica arriba (vacia) + tabla de visitas
    ws = wb.create_sheet(title=codigo[:31])
    ws.cell(1, 1, "FICHA TECNICA").font = neg
    etiquetas = [et for _, et in agente.FICHA_CAMPOS]      # Serie, Marca, Modelo, ...
    r = 2
    for k in range(0, len(etiquetas), 2):
        ws.cell(r, 1, etiquetas[k]).font = neg
        if k + 1 < len(etiquetas):
            ws.cell(r, 4, etiquetas[k + 1]).font = neg
        r += 1
    hr = r + 1                                             # fila en blanco, encabezado abajo
    for j, h in enumerate(_headers(), start=1):
        ws.cell(hr, j, h).font = neg
    dv = DataValidation(type="list", formula1=f'"{CUADRILLAS}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"B{hr + 1}:B{hr + 500}")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = ws.cell(hr + 1, 1)

    # 3) reordenar: Equipos primero, equipos en orden del indice, Limites al final
    orden_idx = [r[1].value for r in indice.iter_rows(min_row=2) if r[1].value]
    orden = ["Equipos"] + orden_idx + ["Limites"]
    orden = [s for s in orden if s in wb.sheetnames]
    wb._sheets.sort(key=lambda s: orden.index(s.title))

    wb.save(archivo)
    print(f"OK: agregado '{codigo}' ({tipo}) en la subestacion '{subestacion}'.")
    print(f"    Ahora abre el Excel y llena las visitas en la hoja '{codigo}'.")
    return True


def _preguntar():
    print("=== Agregar equipo nuevo ===")
    sub = input("Subestacion           : ").strip()
    cod = input("Codigo (ej CCE-32L43) : ").strip()
    tipo = input("Tipo (Interruptor/Restaurador): ").strip()
    return sub, cod, tipo


if __name__ == "__main__":
    if len(sys.argv) == 4:
        sub, cod, tipo = sys.argv[1], sys.argv[2], sys.argv[3]
    elif len(sys.argv) == 1:
        sub, cod, tipo = _preguntar()
    else:
        print('Uso: python agregar_equipo.py "Subestacion" "CODIGO" "Tipo"')
        sys.exit(1)

    if not (sub and cod and tipo):
        print("Faltan datos (subestacion, codigo y tipo son obligatorios).")
        sys.exit(1)

    agregar_equipo(sub, cod, tipo)
